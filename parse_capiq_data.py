"""
Parse CapIQ Income Statement exports (Excel) and compute key financial metrics.

Reads a CapIQ-formatted .xlsx file, extracts the last 12 quarters of Revenue,
Gross Profit, and Operating Income, then calculates margins and YoY changes.
"""

import json
import sys
import glob
import os
from datetime import datetime
import numpy as np
import pandas as pd
import openpyxl


def detect_capiq_units(ws) -> str:
    """Detect the data units from a CapIQ Excel worksheet header.

    Scans the first 25 rows for a cell in column A containing "Units"
    (case-insensitive), then reads the unit value from column B of the
    same row.  CapIQ typically places this in the metadata rows above
    the data table (e.g. row 15), but the scan is intentionally broad
    to handle different header placements.

    Returns:
        Unit string as found in the file (e.g. "Thousands", "Millions").
        Falls back to "Thousands" with a warning if not detected.
    """
    for row_idx in range(1, 26):
        cell_a = ws.cell(row=row_idx, column=1).value
        if cell_a is not None and str(cell_a).strip().lower() == "units":
            cell_b = ws.cell(row=row_idx, column=2).value
            if cell_b is not None:
                return str(cell_b).strip()
    print("  Warning: Could not detect units from CapIQ header; assuming Thousands.")
    return "Thousands"


def capiq_units_to_thousands(units: str) -> float:
    """Return the multiplier to convert from the detected CapIQ units to thousands.

    Examples:
        "Thousands" → 1.0   (already in thousands)
        "Millions"  → 1000.0 (multiply by 1000 to get thousands)
    """
    key = units.lower().strip()
    if key == "thousands":
        return 1.0
    if key == "millions":
        return 1_000.0
    print(f"  Warning: Unrecognized CapIQ unit '{units}'; assuming Thousands.")
    return 1.0


def parse_capiq_income_statement(
    filepath: str, num_quarters: int = 12, debug: bool = False
) -> pd.DataFrame:
    """Parse a CapIQ income statement Excel export and return a metrics DataFrame.

    Args:
        filepath: Path to the CapIQ .xlsx file.
        num_quarters: Number of most-recent quarters to include in output.
        debug: If True, print all row labels found in column A.

    Returns:
        DataFrame with quarterly financial data and computed metrics.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Detect source units and compute normalisation factor to thousands
    detected_units = detect_capiq_units(ws)
    units_multiplier = capiq_units_to_thousands(detected_units)
    print(f"  Income statement units: {detected_units} (multiplier to thousands: {units_multiplier})")

    # --- Label variations for each metric (checked in order, case-insensitive) ---
    metric_variations = {
        "revenue": [
            "Total Revenues",
            "Total Revenue",
            "Net Revenues",
            "Net Revenue",
            "Net Sales",
            "Revenues",
            "Revenue",
            "Ciq Balancing Calc.-revenue",
        ],
        "gross_profit": [
            "Gross Profit/loss",
            "Gross Profit/Loss",
            "Gross Profit",
            "Gross profit",
        ],
        "operating_income": [
            "Operating Income (Loss)",
            "Operating Income",
            "Operating income",
            "Income from Operations",
        ],
        "net_income": [
            "Net Income (Loss)",
            "Net Income",
            "Net Loss",
            "Net Income/Loss",
        ],
    }
    # Build a lookup: lowercased variation -> (metric key, priority)
    # Lower priority number = more specific / preferred match
    variation_lookup = {}
    for key, variations in metric_variations.items():
        for priority, v in enumerate(variations):
            lower = v.strip().lower()
            if lower not in variation_lookup:
                variation_lookup[lower] = (key, priority)

    # --- Locate key rows by scanning column A labels ---
    # Track all candidate matches per metric, then pick the best (lowest priority)
    label_candidates = {}  # metric_key -> list of (priority, row_idx, has_data)
    all_row_labels = []
    period_header_row = None
    period_date_row = None

    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()
        if cell_str:
            all_row_labels.append(cell_str)

        # Period headers row starts with "Recommended:" in col A
        if cell_str.startswith("Recommended:"):
            period_header_row = row_idx
        elif cell_str == "Period Ended":
            period_date_row = row_idx

        match = variation_lookup.get(cell_str.lower())
        if match:
            metric_key, priority = match
            
            # Check if this row has numeric data in any column (not just a section header)
            has_data = False
            for col_idx in range(2, min(ws.max_column + 1, 25)):  # Check first ~20 data columns
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    has_data = True
                    break
            
            label_candidates.setdefault(metric_key, []).append((priority, row_idx, has_data))

    # Pick the best match for each metric
    # Priority: has_data=True > lower priority number > later row
    label_row_map = {}
    for metric_key, candidates in label_candidates.items():
        # Sort by: has_data (True first), priority (lower first), row (later first)
        best = min(candidates, key=lambda x: (not x[2], x[0], -x[1]))
        label_row_map[metric_key] = best[1]

    if debug:
        print(f"\n[DEBUG] All row labels in column A of {os.path.basename(filepath)}:")
        for lbl in all_row_labels:
            print(f"  - {lbl!r}")
        print()

    if period_header_row is None or period_date_row is None:
        raise ValueError("Could not find period header or date rows in the file.")

    # Gross profit and net income are optional; revenue and operating income are required
    missing_optional = {"gross_profit", "net_income"} - set(label_row_map.keys())
    missing_required = {"revenue", "operating_income"} - set(label_row_map.keys())
    warnings = []

    optional_display = {
        "gross_profit": ("Gross Profit", "Gross margin metrics will show as N/A."),
        "net_income": ("Net Income", "ROE and net income metrics will show as N/A."),
    }
    for m in sorted(missing_optional):
        company_name = os.path.basename(filepath).split("_")[1] if "_" in os.path.basename(filepath) else os.path.basename(filepath)
        display_name, note = optional_display[m]
        warnings.append(f"  Warning: {company_name} - {display_name} not found. {note}")

    if missing_required:
        label_names = {
            "revenue": "Revenue",
            "operating_income": "Operating Income",
        }
        msg_parts = [f"Could not find the following metrics in {os.path.basename(filepath)}:"]
        for m in sorted(missing_required):
            tried = ", ".join(f'"{v}"' for v in metric_variations[m])
            msg_parts.append(f"  {label_names[m]} — tried: {tried}")
        msg_parts.append("")
        msg_parts.append("Row labels found in column A:")
        for lbl in all_row_labels:
            msg_parts.append(f"  - {lbl!r}")
        raise ValueError("\n".join(msg_parts))

    for w in warnings:
        print(w)

    # --- Read period headers and dates from data columns (col 2 onward) ---
    quarters = []
    dates = []
    data_cols = []
    for col_idx in range(2, ws.max_column + 1):
        header = ws.cell(row=period_header_row, column=col_idx).value
        if header is None:
            continue
        quarters.append(str(header).strip())
        date_val = ws.cell(row=period_date_row, column=col_idx).value
        dates.append(pd.Timestamp(date_val) if date_val else None)
        data_cols.append(col_idx)

    total_available = len(quarters)
    # We need extra quarters before the output window to compute YoY (4 prior quarters)
    yoy_lookback = 4
    if total_available < num_quarters:
        print(
            f"  Note: file has {total_available} quarters "
            f"(requested {num_quarters}), using all available."
        )
        num_quarters = total_available
    needed = num_quarters + yoy_lookback

    # Determine the slice: take as many as we can for YoY, at minimum num_quarters
    start_idx = max(0, total_available - needed)
    end_idx = total_available

    def read_row_values(row_num):
        """Read numeric values for data columns, applying units_multiplier to normalize to thousands."""
        vals = []
        for col_idx in data_cols[start_idx:end_idx]:
            raw = ws.cell(row=row_num, column=col_idx).value
            if raw is None or str(raw).strip().upper() == "NA":
                vals.append(None)
            else:
                vals.append(float(raw) * units_multiplier)
        return vals

    revenue = read_row_values(label_row_map["revenue"])
    if "gross_profit" in label_row_map:
        gross_profit = read_row_values(label_row_map["gross_profit"])
    else:
        gross_profit = [None] * (end_idx - start_idx)
    operating_income = read_row_values(label_row_map["operating_income"])
    if "net_income" in label_row_map:
        net_income = read_row_values(label_row_map["net_income"])
    else:
        net_income = [None] * (end_idx - start_idx)
    q_labels = quarters[start_idx:end_idx]
    q_dates = dates[start_idx:end_idx]

    # --- Build full working DataFrame (may be wider than output for YoY calc) ---
    work = pd.DataFrame({
        "Quarter": q_labels,
        "Date": q_dates,
        "Revenue": revenue,
        "Gross_Profit": gross_profit,
        "Operating_Income": operating_income,
        "Net_Income": net_income,
    })

    work["Gross_Margin"] = work["Gross_Profit"] / work["Revenue"]
    work["Operating_Margin"] = work["Operating_Income"] / work["Revenue"]

    # YoY calculations: compare index i to i-4 (same fiscal quarter, prior year)
    work["Revenue_Growth_YoY"] = None
    work["Gross_Margin_Delta_YoY"] = None
    work["Operating_Margin_Delta_YoY"] = None

    for i in range(yoy_lookback, len(work)):
        prev = i - yoy_lookback
        if work.loc[prev, "Revenue"] and work.loc[prev, "Revenue"] != 0:
            work.loc[i, "Revenue_Growth_YoY"] = (
                (work.loc[i, "Revenue"] - work.loc[prev, "Revenue"])
                / abs(work.loc[prev, "Revenue"])
            )
        if work.loc[prev, "Gross_Margin"] is not None:
            work.loc[i, "Gross_Margin_Delta_YoY"] = (
                work.loc[i, "Gross_Margin"] - work.loc[prev, "Gross_Margin"]
            )
        if work.loc[prev, "Operating_Margin"] is not None:
            work.loc[i, "Operating_Margin_Delta_YoY"] = (
                work.loc[i, "Operating_Margin"] - work.loc[prev, "Operating_Margin"]
            )

    # Trim to the last num_quarters rows (the output window)
    result = work.tail(num_quarters).reset_index(drop=True)

    # Format Date column
    result["Date"] = result["Date"].dt.strftime("%Y-%m-%d")

    return result


def parse_capiq_balance_sheet(
    filepath: str, num_quarters: int = 12, debug: bool = False
) -> pd.DataFrame:
    """Parse a CapIQ balance sheet Excel export and return a DataFrame.

    Extracts Cash, Total Debt, Total Equity, Current Assets, Current Liabilities,
    and Total Assets for each quarter.  Total Debt is computed as the sum of all
    debt line-items found (e.g. long-term debt, convertible notes).

    Args:
        filepath: Path to the CapIQ balance sheet .xlsx file.
        num_quarters: Number of most-recent quarters to include in output.
        debug: If True, print all row labels found in column A.

    Returns:
        DataFrame with Quarter, Date, and the six balance-sheet columns.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Detect source units and compute normalisation factor to thousands
    detected_units = detect_capiq_units(ws)
    units_multiplier = capiq_units_to_thousands(detected_units)
    print(f"  Balance sheet units: {detected_units} (multiplier to thousands: {units_multiplier})")

    # --- Label variations for single-row metrics (checked in order, case-insensitive) ---
    metric_variations = {
        "cash": [
            "Cash and Cash Equivalents",
            "Cash & Cash Equivalents",
            "Cash & Equivalents",
            "Cash and Equivalents",
            "Cash",
        ],
        "total_equity": [
            "Total Shareholders Equity",
            "Total Stockholders' Equity",
            "Total Shareholders' Equity",
            "Total Stockholders Equity",
            "Total Equity",
        ],
        "current_assets": [
            "Total Current Assets",
        ],
        "current_liabilities": [
            "Total Current Liabilities",
        ],
        "total_assets": [
            "Total Assets",
        ],
    }

    # Debt component labels — ALL matching rows are summed to produce Total Debt
    debt_labels = [
        "Convertible Senior Notes, Net Current",
        "Convertible Senior Notes, Net",
        "Long-term Debt",
        "Long-term Debt Net",
        "Long-term Debt, Net",
        "Long-term Debts, Net of Current Portion",
        "Long-term Borrowings",
        "Notes Payable and Long-term Debt, Less Current Installments",
        "Notes Payable, Long-term Debt and Long-term Lease Obligations Less Current Installments",
        "Long-term Portion of Borrowings Under Credit Facility and Lease Obligations",
        "Short-term Borrowings",
        "Short-term Debt",
        "Borrowings, Current",
        "Bank Borrowings and Current Portion of Long-term Debt",
        "Current Portion of Long-term Debt",
        "Current Portion Of Long-Term Debt",
        "Current Installments of Notes Payable and Long-term Debt",
        "Current Installments of Notes Payable, Long-term Debt and Capital Lease Obligations",
        "Current Portion of Borrowings Under Credit Facility and Finance Lease Obligations",
    ]

    # Deferred revenue component labels — ALL matching rows are summed
    deferred_revenue_labels = [
        "Deferred Revenue",
        "Deferred Revenues",
        "Deferred Revenue-current",
        "Deferred Revenue-noncurrent",
        "Deferred Revenue, Current",
        "Deferred Revenue, Non-Current",
        "Contract Liabilities",
        "Unearned Revenue",
    ]

    # Build lookup: lowercased variation -> (metric key, priority)
    variation_lookup = {}
    for key, variations in metric_variations.items():
        for priority, v in enumerate(variations):
            lower = v.strip().lower()
            if lower not in variation_lookup:
                variation_lookup[lower] = (key, priority)

    debt_label_set = {lbl.strip().lower() for lbl in debt_labels}
    deferred_rev_label_set = {lbl.strip().lower() for lbl in deferred_revenue_labels}

    # --- Scan column A for row labels ---
    label_candidates = {}  # metric_key -> list of (priority, row_idx, has_data)
    debt_rows = []  # row indices containing debt data to sum
    deferred_rev_rows = []  # row indices containing deferred revenue data to sum
    all_row_labels = []
    period_header_row = None
    period_date_row = None

    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()
        if cell_str:
            all_row_labels.append(cell_str)

        if cell_str.startswith("Recommended:"):
            period_header_row = row_idx
        elif cell_str == "Period Ended":
            period_date_row = row_idx

        cell_lower = cell_str.lower()

        # Check single-row metric match
        match = variation_lookup.get(cell_lower)
        if match:
            metric_key, priority = match
            has_data = False
            for col_idx in range(2, min(ws.max_column + 1, 25)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    has_data = True
                    break
            label_candidates.setdefault(metric_key, []).append(
                (priority, row_idx, has_data)
            )

        # Check debt component match
        if cell_lower in debt_label_set:
            has_data = False
            for col_idx in range(2, min(ws.max_column + 1, 25)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    has_data = True
                    break
            if has_data:
                debt_rows.append(row_idx)

        # Check deferred revenue component match
        if cell_lower in deferred_rev_label_set:
            has_data = False
            for col_idx in range(2, min(ws.max_column + 1, 25)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    has_data = True
                    break
            if has_data:
                deferred_rev_rows.append(row_idx)

    # Pick best match per single-row metric
    label_row_map = {}
    for metric_key, candidates in label_candidates.items():
        best = min(candidates, key=lambda x: (not x[2], x[0], -x[1]))
        label_row_map[metric_key] = best[1]

    if debug:
        print(f"\n[DEBUG] All row labels in column A of {os.path.basename(filepath)}:")
        for lbl in all_row_labels:
            print(f"  - {lbl!r}")
        print(f"\n[DEBUG] Single-row matches: {label_row_map}")
        print(f"[DEBUG] Debt component rows: {debt_rows}")
        print(f"[DEBUG] Deferred revenue component rows: {deferred_rev_rows}")
        print()

    if period_header_row is None or period_date_row is None:
        raise ValueError("Could not find period header or date rows in the file.")

    # Validate required single-row metrics (deferred_revenue is optional)
    required = {"cash", "total_equity", "current_assets", "current_liabilities", "total_assets"}
    missing = required - set(label_row_map.keys())
    if missing:
        label_names = {
            "cash": "Cash",
            "total_equity": "Total Equity",
            "current_assets": "Total Current Assets",
            "current_liabilities": "Total Current Liabilities",
            "total_assets": "Total Assets",
        }
        msg_parts = [f"Could not find the following metrics in {os.path.basename(filepath)}:"]
        for m in sorted(missing):
            tried = ", ".join(f'"{v}"' for v in metric_variations[m])
            msg_parts.append(f"  {label_names[m]} — tried: {tried}")
        msg_parts.append("")
        msg_parts.append("Row labels found in column A:")
        for lbl in all_row_labels:
            msg_parts.append(f"  - {lbl!r}")
        raise ValueError("\n".join(msg_parts))

    if not deferred_rev_rows:
        company_name = (
            os.path.basename(filepath).split("_")[1]
            if "_" in os.path.basename(filepath)
            else os.path.basename(filepath)
        )
        print(
            f"  Warning: {company_name} - No deferred revenue line items found. "
            "Deferred revenue metrics will show as N/A."
        )

    if not debt_rows:
        company_name = (
            os.path.basename(filepath).split("_")[1]
            if "_" in os.path.basename(filepath)
            else os.path.basename(filepath)
        )
        print(
            f"  Warning: {company_name} - No debt line items found. "
            "Total Debt will show as 0."
        )

    # --- Read period headers and dates ---
    quarters = []
    dates = []
    data_cols = []
    for col_idx in range(2, ws.max_column + 1):
        header = ws.cell(row=period_header_row, column=col_idx).value
        if header is None:
            continue
        quarters.append(str(header).strip())
        date_val = ws.cell(row=period_date_row, column=col_idx).value
        dates.append(pd.Timestamp(date_val) if date_val else None)
        data_cols.append(col_idx)

    total_available = len(quarters)
    if total_available < num_quarters:
        print(
            f"  Note: file has {total_available} quarters "
            f"(requested {num_quarters}), using all available."
        )
        num_quarters = total_available

    start_idx = max(0, total_available - num_quarters)
    end_idx = total_available

    def read_row_values(row_num):
        """Read numeric values for data columns, applying units_multiplier to normalize to thousands."""
        vals = []
        for col_idx in data_cols[start_idx:end_idx]:
            raw = ws.cell(row=row_num, column=col_idx).value
            if raw is None or str(raw).strip().upper() == "NA":
                vals.append(None)
            else:
                vals.append(float(raw) * units_multiplier)
        return vals

    # Read single-row metrics
    cash = read_row_values(label_row_map["cash"])
    total_equity = read_row_values(label_row_map["total_equity"])
    current_assets = read_row_values(label_row_map["current_assets"])
    current_liabilities = read_row_values(label_row_map["current_liabilities"])
    total_assets = read_row_values(label_row_map["total_assets"])

    # Compute deferred revenue by summing all component rows (current + non-current)
    n = end_idx - start_idx
    if deferred_rev_rows:
        deferred_revenue = [0.0] * n
        for dr_row in deferred_rev_rows:
            vals = read_row_values(dr_row)
            for i in range(n):
                if vals[i] is not None:
                    deferred_revenue[i] += vals[i]
    else:
        deferred_revenue = [None] * n

    # Compute total debt by summing all debt component rows
    if debt_rows:
        total_debt = [0.0] * n
        for dr in debt_rows:
            vals = read_row_values(dr)
            for i in range(n):
                if vals[i] is not None:
                    total_debt[i] += vals[i]
    else:
        total_debt = [0.0] * n

    q_labels = quarters[start_idx:end_idx]
    q_dates = dates[start_idx:end_idx]

    result = pd.DataFrame({
        "Quarter": q_labels,
        "Date": [d.strftime("%Y-%m-%d") if d else None for d in q_dates],
        "Cash": cash,
        "Total_Debt": total_debt,
        "Total_Equity": total_equity,
        "Current_Assets": current_assets,
        "Current_Liabilities": current_liabilities,
        "Total_Assets": total_assets,
        "Deferred_Revenue": deferred_revenue,
    })

    return result


def parse_capiq_cashflow(
    filepath: str, num_quarters: int = 12, debug: bool = False
) -> pd.DataFrame:
    """Parse a CapIQ cash flow statement Excel export and return a DataFrame.

    Extracts Operating Cash Flow and CapEx for each quarter.

    Args:
        filepath: Path to the CapIQ cash flow .xlsx file.
        num_quarters: Number of most-recent quarters to include in output.
        debug: If True, print all row labels found in column A.

    Returns:
        DataFrame with Quarter, Date, Operating_Cash_Flow, CapEx.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Detect source units and compute normalisation factor to thousands
    detected_units = detect_capiq_units(ws)
    units_multiplier = capiq_units_to_thousands(detected_units)

    metric_variations = {
        "operating_cash_flow": [
            "Cash Flow from Operating Activities",
            "Operating Cash Flow",
            "Net Cash from Operating Activities",
            "Cash from Operations",
        ],
        "capex": [
            "Purchase of Computer Software and Property, Plant and Equipment",
            "Purchase of Capital Assets",
            "Acquisition of Property, Plant and Equipment",
            "Purchase of Property and Equipment",
            "Purchase of Property, Equipment and Other Assets",
            "Capital Expenditures",
            "Capex",
        ],
    }

    variation_lookup = {}
    for key, variations in metric_variations.items():
        for priority, v in enumerate(variations):
            lower = v.strip().lower()
            if lower not in variation_lookup:
                variation_lookup[lower] = (key, priority)

    label_candidates = {}
    all_row_labels = []
    period_header_row = None
    period_date_row = None

    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()
        if cell_str:
            all_row_labels.append(cell_str)

        if cell_str.startswith("Recommended:"):
            period_header_row = row_idx
        elif cell_str == "Period Ended":
            period_date_row = row_idx

        match = variation_lookup.get(cell_str.lower())
        if match:
            metric_key, priority = match
            has_data = False
            for col_idx in range(2, min(ws.max_column + 1, 25)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, (int, float)):
                    has_data = True
                    break
            label_candidates.setdefault(metric_key, []).append(
                (priority, row_idx, has_data)
            )

    label_row_map = {}
    for metric_key, candidates in label_candidates.items():
        best = min(candidates, key=lambda x: (not x[2], x[0], -x[1]))
        label_row_map[metric_key] = best[1]

    if debug:
        print(f"\n[DEBUG] All row labels in column A of {os.path.basename(filepath)}:")
        for lbl in all_row_labels:
            print(f"  - {lbl!r}")
        print(f"\n[DEBUG] Matched rows: {label_row_map}")
        print()

    if period_header_row is None or period_date_row is None:
        raise ValueError("Could not find period header or date rows in the file.")

    required = {"operating_cash_flow", "capex"}
    missing = required - set(label_row_map.keys())
    if missing:
        label_names = {
            "operating_cash_flow": "Operating Cash Flow",
            "capex": "CapEx",
        }
        msg_parts = [f"Could not find the following metrics in {os.path.basename(filepath)}:"]
        for m in sorted(missing):
            tried = ", ".join(f'"{v}"' for v in metric_variations[m])
            msg_parts.append(f"  {label_names[m]} — tried: {tried}")
        msg_parts.append("")
        msg_parts.append("Row labels found in column A:")
        for lbl in all_row_labels:
            msg_parts.append(f"  - {lbl!r}")
        raise ValueError("\n".join(msg_parts))

    quarters = []
    dates = []
    data_cols = []
    for col_idx in range(2, ws.max_column + 1):
        header = ws.cell(row=period_header_row, column=col_idx).value
        if header is None:
            continue
        quarters.append(str(header).strip())
        date_val = ws.cell(row=period_date_row, column=col_idx).value
        dates.append(pd.Timestamp(date_val) if date_val else None)
        data_cols.append(col_idx)

    total_available = len(quarters)
    if total_available < num_quarters:
        print(
            f"  Note: file has {total_available} quarters "
            f"(requested {num_quarters}), using all available."
        )
        num_quarters = total_available

    start_idx = max(0, total_available - num_quarters)
    end_idx = total_available

    def read_row_values(row_num):
        vals = []
        for col_idx in data_cols[start_idx:end_idx]:
            raw = ws.cell(row=row_num, column=col_idx).value
            if raw is None or str(raw).strip().upper() == "NA":
                vals.append(None)
            else:
                vals.append(float(raw) * units_multiplier)
        return vals

    ocf = read_row_values(label_row_map["operating_cash_flow"])
    capex = read_row_values(label_row_map["capex"])

    q_labels = quarters[start_idx:end_idx]
    q_dates = dates[start_idx:end_idx]

    result = pd.DataFrame({
        "Quarter": q_labels,
        "Date": [d.strftime("%Y-%m-%d") if d else None for d in q_dates],
        "Operating_Cash_Flow": ocf,
        "CapEx": capex,
    })

    return result


def calculate_ttm_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Calculate trailing-twelve-month (TTM) metrics for each quarter.
    
    For each quarter, sums the last 4 quarters of revenue/profit metrics
    and calculates TTM margins. This smooths quarterly volatility and is
    particularly useful for seasonal businesses.
    
    Args:
        df: DataFrame with quarterly data including Revenue, Gross_Profit, Operating_Income
        
    Returns:
        DataFrame with additional TTM columns:
        - TTM_Revenue: Sum of last 4 quarters revenue
        - TTM_Gross_Profit: Sum of last 4 quarters gross profit
        - TTM_Operating_Income: Sum of last 4 quarters operating income
        - TTM_Gross_Margin: TTM gross profit / TTM revenue
        - TTM_Operating_Margin: TTM operating income / TTM revenue
        - TTM_Net_Income: Sum of last 4 quarters net income (if available)
    """
    df = df.copy()
    
    # Calculate rolling 4-quarter sums
    df["TTM_Revenue"] = df["Revenue"].rolling(window=4, min_periods=4).sum()
    df["TTM_Gross_Profit"] = df["Gross_Profit"].rolling(window=4, min_periods=4).sum()
    df["TTM_Operating_Income"] = df["Operating_Income"].rolling(window=4, min_periods=4).sum()
    
    # Calculate TTM margins
    df["TTM_Gross_Margin"] = df["TTM_Gross_Profit"] / df["TTM_Revenue"].replace(0, np.nan)
    df["TTM_Operating_Margin"] = df["TTM_Operating_Income"] / df["TTM_Revenue"].replace(0, np.nan)
    
    # Calculate TTM Net Income if available
    if "Net_Income" in df.columns:
        df["TTM_Net_Income"] = df["Net_Income"].rolling(window=4, min_periods=4).sum()
    
    return df


def main():
    debug = "--debug" in sys.argv
    args = [a for a in sys.argv[1:] if a != "--debug"]

    script_dir = os.path.dirname(os.path.abspath(__file__))
    export_dir = os.path.join(script_dir, "CapIQ Exports")

    if args:
        files = args
    else:
        # Default: look for income statement files in CapIQ Exports folder
        pattern = os.path.join(export_dir, "*IncomeStatement*")
        files = sorted(glob.glob(pattern))
        if not files:
            print("No CapIQ income statement files found. Pass file paths as arguments.")
            sys.exit(1)

    # Build lookups of available balance sheet and cash flow files by company name
    bs_files_by_company = {}
    for bs_path in glob.glob(os.path.join(export_dir, "*BalanceSheet*")):
        bs_name = os.path.basename(bs_path)
        bs_company = bs_name.split("_")[1] if "_" in bs_name else None
        if bs_company:
            bs_files_by_company[bs_company] = bs_path

    cf_files_by_company = {}
    for cf_path in glob.glob(os.path.join(export_dir, "*CashFlow*")):
        cf_name = os.path.basename(cf_path)
        cf_company = cf_name.split("_")[1] if "_" in cf_name else None
        if cf_company:
            cf_files_by_company[cf_company] = cf_path

    pd.set_option("display.float_format", lambda x: f"{x:,.2f}")
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 200)

    for filepath in files:
        filename = os.path.basename(filepath)
        company = filename.split("_")[1] if "_" in filename else filename
        print(f"\n{'='*80}")
        print(f"Company: {company}")
        print(f"File:    {filename}")
        print(f"{'='*80}\n")

        df = parse_capiq_income_statement(filepath, debug=debug)

        # Try to find and merge a matching balance sheet
        bs_path = bs_files_by_company.get(company)
        if bs_path:
            bs_df = parse_capiq_balance_sheet(bs_path, debug=debug)
            df = df.merge(bs_df, on="Quarter", how="left", suffixes=("", "_bs"))
            # Drop the duplicate Date column from balance sheet
            if "Date_bs" in df.columns:
                df.drop(columns=["Date_bs"], inplace=True)
            print(f"  Merged balance sheet: {os.path.basename(bs_path)}")

            # Calculate financial ratios from merged data
            df["Net_Debt"] = df["Total_Debt"] - df["Cash"]
            df["Debt_to_Equity"] = df["Total_Debt"] / df["Total_Equity"].replace(0, np.nan)
            df["Current_Ratio"] = df["Current_Assets"] / df["Current_Liabilities"].replace(0, np.nan)
            df["ROE"] = df["Net_Income"] / df["Total_Equity"].replace(0, np.nan)
            df["Asset_Turnover"] = df["Revenue"] / df["Total_Assets"].replace(0, np.nan)

            # Deferred revenue metrics
            df["Revenue_Recognition_Quality"] = df["Deferred_Revenue"] / df["Revenue"].replace(0, np.nan)

            # Deferred Revenue Growth YoY: compare quarter i to quarter i-4
            df["Deferred_Revenue_Growth_YoY"] = None
            for i in range(4, len(df)):
                prev_dr = df.loc[df.index[i - 4], "Deferred_Revenue"]
                curr_dr = df.loc[df.index[i], "Deferred_Revenue"]
                if pd.notna(prev_dr) and pd.notna(curr_dr) and prev_dr != 0:
                    df.loc[df.index[i], "Deferred_Revenue_Growth_YoY"] = (
                        (curr_dr - prev_dr) / abs(prev_dr)
                    )

            # Replace inf/-inf from edge cases (e.g. near-zero denominators)
            ratio_cols = [
                "Net_Debt", "Debt_to_Equity", "Current_Ratio", "ROE", "Asset_Turnover",
                "Revenue_Recognition_Quality", "Deferred_Revenue_Growth_YoY",
            ]
            df[ratio_cols] = df[ratio_cols].replace([np.inf, -np.inf], np.nan)

            # --- Sanity check: deferred revenue vs revenue ---
            dr_vals = df["Deferred_Revenue"]
            if dr_vals.notna().any():
                first_valid = dr_vals.first_valid_index()
                last_valid = dr_vals.last_valid_index()
                first_row = df.loc[first_valid]
                last_row = df.loc[last_valid]
                print(f"\n  Deferred Revenue sanity check (all values in $K):")
                print(f"    First quarter with data: {first_row['Quarter']}")
                print(f"      Deferred Revenue (raw $K): {first_row['Deferred_Revenue']:>12,.0f}")
                print(f"      Revenue          (raw $K): {first_row['Revenue']:>12,.0f}")
                print(f"      DR / Revenue ratio:        {first_row['Deferred_Revenue'] / first_row['Revenue']:>12.2f}x")
                print(f"      Deferred Revenue ($M):     {first_row['Deferred_Revenue'] / 1_000:>12,.1f}")
                print(f"      Revenue          ($M):     {first_row['Revenue'] / 1_000:>12,.1f}")
                print(f"    Last quarter with data:  {last_row['Quarter']}")
                print(f"      Deferred Revenue (raw $K): {last_row['Deferred_Revenue']:>12,.0f}")
                print(f"      Revenue          (raw $K): {last_row['Revenue']:>12,.0f}")
                print(f"      DR / Revenue ratio:        {last_row['Deferred_Revenue'] / last_row['Revenue']:>12.2f}x")
                print(f"      Deferred Revenue ($M):     {last_row['Deferred_Revenue'] / 1_000:>12,.1f}")
                print(f"      Revenue          ($M):     {last_row['Revenue'] / 1_000:>12,.1f}")
        else:
            print(f"  Note: No balance sheet found for {company}")

        # Calculate TTM (trailing-twelve-month) metrics
        df = calculate_ttm_metrics(df)
        print(f"  Calculated TTM metrics (requires 4+ quarters of data)")

        print(df.to_string(index=False))

        # Save quarterly CSV to processed_data subdirectory
        out_dir = os.path.join(script_dir, "processed_data")
        os.makedirs(out_dir, exist_ok=True)
        out_name = f"{company}_quarterly_metrics.csv"
        out_path = os.path.join(out_dir, out_name)
        df.to_csv(out_path, index=False)
        print(f"\nSaved to: {out_path}")

        # Save metadata alongside CSV
        metadata = {
            "processed_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_file": os.path.basename(filepath),
            "latest_quarter": str(df.iloc[-1]["Quarter"]),
            "latest_date": str(df.iloc[-1]["Date"])[:10],
        }
        metadata_path = out_path.replace(".csv", "_metadata.json")
        with open(metadata_path, "w") as f:
            json.dump(metadata, f, indent=2)

        # --- Annual cash flow (separate file) ---
        cf_path = cf_files_by_company.get(company)
        if cf_path:
            cf_df = parse_capiq_cashflow(cf_path, debug=debug)
            # FCF = OCF + CapEx (CapEx is negative in the data, so adding subtracts)
            cf_df["Free_Cash_Flow"] = cf_df["Operating_Cash_Flow"] + cf_df["CapEx"]

            # Compute annual revenue by summing matching fiscal quarters
            annual_rev = []
            for _, cf_row in cf_df.iterrows():
                year_prefix = cf_row["Quarter"].split()[0]  # "2025 FY" -> "2025"
                q_mask = df["Quarter"].str.startswith(year_prefix + " FQ")
                matching = df.loc[q_mask, "Revenue"]
                if len(matching) == 4:
                    annual_rev.append(matching.sum())
                else:
                    annual_rev.append(None)
            cf_df["Annual_Revenue"] = annual_rev
            cf_df["FCF_Margin"] = cf_df["Free_Cash_Flow"] / cf_df["Annual_Revenue"]

            cf_out = os.path.join(out_dir, f"{company}_annual_cashflow.csv")
            cf_df.to_csv(cf_out, index=False)
            print(f"  Saved annual cash flow to: {cf_out}")
        else:
            print(f"  Note: No cash flow file found for {company}")


if __name__ == "__main__":
    main()
